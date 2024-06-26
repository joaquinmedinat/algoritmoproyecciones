# -*- coding: utf-8 -*-
"""AlgoritmoProyecciones.ipynb


"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl import cell
import copy


wb=Workbook()
ws=wb.active
def encabezados(dataf,fila):
  e=list(dataf.columns)
  for i in range(1,len(e)+1):
    ws.cell(fila,i).value=e[i-1]

def transcribirToExcel(datadic,dicDatosContratos):
  filaEscribir=1

  ws.cell(filaEscribir,1).value="centro"
  ws.cell(filaEscribir,2).value="contrato"
  ws.cell(filaEscribir,3).value="tasa"
  ws.cell(filaEscribir,4).value="periocidad"

  ws.cell(filaEscribir,5).value="Proyeccion 2024"
  ws.cell(filaEscribir,6).value="Proyeccion 2025"
  ws.cell(filaEscribir,7).value="Proyeccion 2026"

  for key in datadic:

    filaEscribir+=1

    data=datadic[key]
    dContratoActual=dicDatosContratos[key]

    ws.cell(filaEscribir,1).value=dContratoActual["centro"]
    ws.cell(filaEscribir,2).value=dContratoActual["contrato"]
    ws.cell(filaEscribir,3).value=dContratoActual["tasa"]
    ws.cell(filaEscribir,4).value=dContratoActual["periocidad"]
    ws.cell(filaEscribir,5).value=dContratoActual["2024"]
    ws.cell(filaEscribir,6).value=dContratoActual["2025"]
    ws.cell(filaEscribir,7).value=dContratoActual["2026"]



    filaEscribir+=1
    #lo escribe en la filan num 2.
    encabezados(data,filaEscribir)
    for numFila,fila in data.iterrows():
      filaEscribir+=1
      #print(numFila)
      for numCol in range(len(fila)):
        #print(numFila,numCol)
        #print(filaEscribir,numCol)
        ws.cell(filaEscribir,numCol+1).value=data.iat[numFila-1,numCol]
    filaEscribir+=1

  wb.save("Proyecciones.xlsx")

def calcularReajuste(tasa,precio):
  t=1+tasa/100

  p=precio*t

  return p




# Change to the /content directory
#os.chdir('/content')
os.chdir('/content')
#wb = load_workbook(filename = 'test.xlsx')
#hojas=wb.sheetnames
#hojas.remove("centros")

excel_file="test.xlsx"
sheet_name = 'Hoja1'
excel_data = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

datosContratos= pd.read_excel("datoscontratos.xlsx", sheet_name="datosContratos")
ordenDatos=datosContratos.pop(0)
datosContratos.index=ordenDatos


tarifarios={}

formarTabla=[]
nombreTarifario=False

c=["Item","Precio","Cantidad"]

for index, row in excel_data.iterrows():
    if index==0 or nombreTarifario:
        contrato=row[0]+"-"+str(row[1])
        nombreTarifario=False
        #print(contrato)
        #continue

    elif row.isnull().all() or row[0]=="*":
        formarTabla.pop(0)
        cantFilas=len(formarTabla)
        formarTabla=pd.DataFrame(formarTabla)
        formarTabla.columns=c
        formarTabla.index=[i for i in range(1,cantFilas+1)]
        #tarifarios.append(formarTabla)
        tarifarios[contrato]=formarTabla
        formarTabla=[]
        nombreTarifario=True

    else:
        formarTabla.append(row)

tarifariosSf=copy.deepcopy(tarifarios)
tarifariosCf=copy.deepcopy(tarifarios)

b=False

#tarifariosSf[t][""]
dicDatosContratos={}
for t in tarifariosSf:
    subDic={}

    cwyCentro=t.split("-")
    contrato=cwyCentro[0]
    centro=cwyCentro[1]
    tasa=datosContratos["tasa"][contrato]
    periocidad=datosContratos["periocidad"][contrato]

    subDic["contrato"]=contrato
    subDic["centro"]=centro
    subDic["tasa"]=tasa
    subDic["periocidad"]=periocidad


    #tarifariosSf[t][""]

    numReajustesYear=int(12/periocidad)
    nTotalReajustes=numReajustesYear*3
    contador=0
    year=2024

    tarifariosSf[t]["Costo mensual tarifario actual"]=tarifariosSf[t]["Precio"]*tarifariosSf[t]["Cantidad"]
    tarifariosSf[t][f"Costo periodo {year}-1"]=tarifariosSf[t]["Costo mensual tarifario actual"]*periocidad

    dicCostosAños={"2024":0,"2025":0,"2026":0}

    dicCostosAños[f"{year}"]+=tarifariosSf[t][f"Costo periodo {year}-1"].sum()

    price=tarifariosSf[t]["Precio"]
    for i in range(nTotalReajustes):
          #print("entre")
          if year!=2024 or i!=0:
            #print("r")
            tarifariosSf[t][f"Reajuste {year}-{contador+1}"]=calcularReajuste(tasa,price)
            tarifariosSf[t][f"Costo periodo {year}-{contador+1}"]=tarifariosSf[t][f"Reajuste {year}-{contador+1}"]*periocidad*tarifariosSf[t]["Cantidad"]
            dicCostosAños[f"{year}"]+=tarifariosSf[t][f"Costo periodo {year}-{contador+1}"].sum()
            price=tarifariosSf[t][f"Reajuste {year}-{contador+1}"]
          contador+=1
          if contador==numReajustesYear:
            year+=1
            contador=0

    subDic["2024"]=dicCostosAños['2024']
    subDic["2025"]=dicCostosAños['2025']
    subDic["2026"]=dicCostosAños['2026']

    dicDatosContratos[t]=subDic

#for t in tarifariosCf:
#    tarifariosSf[t]["costoDiciembre"]=f"{tarifariosSf[t]["Precio"]}*tarifariosSf[t]["Cantidad"]"

transcribirToExcel(tarifariosSf,dicDatosContratos)
